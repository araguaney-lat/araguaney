"""El anexo Carta Porte como hoja de cálculo.

La versión JSON es para quien integra con su PAC; esta es para quien la abre,
la revisa y la manda por correo. Por eso lo que falta va arriba y en rojo: es
lo primero que tiene que ver quien la recibe, no una nota al pie.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_COLS = (
    ("Descripción", 42),
    ("ClaveProdServCP", 18),
    ("Cantidad", 12),
    ("ClaveUnidad", 14),
    ("Unidad", 16),
    ("PesoEnKg", 14),
)

_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_ALERT_FILL = PatternFill("solid", fgColor="FDE8E8")


def build_carta_porte_xlsx(anexo: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Carta Porte 3.1"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"ANEXO CARTA PORTE 3.1 · {anexo.get('Referencia') or '—'}"
    ws["A1"].font = Font(bold=True, size=13)

    ws.merge_cells("A2:F2")
    ws["A2"] = anexo["_aviso"]
    ws["A2"].font = Font(size=8, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32

    fila = 4

    # Lo que falta va antes que los datos: quien abre esto necesita saber qué
    # completar antes de leer nada más.
    if anexo["_faltantes"]:
        ws.merge_cells(f"A{fila}:F{fila}")
        celda = ws.cell(row=fila, column=1, value="FALTA PARA PODER TIMBRAR")
        celda.font = Font(bold=True, size=10)
        celda.fill = _ALERT_FILL
        fila += 1
        for texto in anexo["_faltantes"]:
            ws.merge_cells(f"A{fila}:F{fila}")
            c = ws.cell(row=fila, column=1, value=f"· {texto}")
            c.font = Font(size=9)
            c.fill = _ALERT_FILL
            c.alignment = Alignment(wrap_text=True)
            fila += 1
        fila += 1

    for etiqueta, clave in (
        ("Origen", "Origen"), ("Destino", "Destino"),
        ("Peso bruto total", "PesoBrutoTotal"), ("Unidad de peso", "UnidadPeso"),
        ("Número de bultos", "NumeroBultos"), ("Total de mercancías", "NumTotalMercancias"),
    ):
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True, size=9)
        ws.cell(row=fila, column=2, value=anexo.get(clave) if anexo.get(clave) is not None else "—")
        fila += 1

    fila += 1
    for i, (nombre, ancho) in enumerate(_COLS, start=1):
        c = ws.cell(row=fila, column=i, value=nombre)
        c.font = Font(bold=True, size=9)
        c.fill = _HEADER_FILL
        ws.column_dimensions[c.column_letter].width = ancho
    fila += 1

    for m in anexo["mercancias"]:
        for i, clave in enumerate(
            ("Descripcion", "ClaveProdServCP", "Cantidad", "ClaveUnidad", "Unidad", "PesoEnKg"),
            start=1,
        ):
            valor = m.get(clave)
            c = ws.cell(row=fila, column=i, value=valor if valor is not None else "—")
            c.font = Font(size=9)
            # La celda vacía se ve: es la que alguien tiene que llenar.
            if valor is None:
                c.fill = _ALERT_FILL
        fila += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
