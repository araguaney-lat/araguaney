"""La declaración de mercancías como hoja de cálculo.

La versión JSON es para quien integra con un sistema; esta es para quien la
abre, la revisa y la manda por correo. Por eso lo que falta va arriba y
resaltado: es lo primero que tiene que ver quien la recibe, no una nota al pie.

Funciona igual con o sin perfil de país — los campos universales siempre están,
y el perfil solo agrega los nombres traducidos.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_COLS = (
    ("Descripción", 42),
    ("Código HS", 16),
    ("Cantidad", 12),
    ("Unidad", 16),
    ("Peso kg", 14),
)
_FIELDS = ("description", "hs_code", "quantity", "unit", "weight_kg")

_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_ALERT_FILL = PatternFill("solid", fgColor="FDE8E8")


def build_declaration_xlsx(doc: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Declaración de mercancías"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"DECLARACIÓN DE MERCANCÍAS · {doc.get('reference') or '—'}"
    ws["A1"].font = Font(bold=True, size=13)

    ws.merge_cells("A2:E2")
    ws["A2"] = doc["_notice"]
    ws["A2"].font = Font(size=8, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32

    row = 4

    # Lo que falta va antes que los datos: quien abre esto necesita saber qué
    # completar antes de leer nada más.
    if doc["_missing"]:
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws.cell(row=row, column=1, value="DATOS QUE FALTAN")
        cell.font = Font(bold=True, size=10)
        cell.fill = _ALERT_FILL
        row += 1
        for text in doc["_missing"]:
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row=row, column=1, value=f"· {text}")
            c.font = Font(size=9)
            c.fill = _ALERT_FILL
            c.alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    issuer = doc.get("issuer") or {}
    for label, value in (
        ("Emisor (razón social)", issuer.get("legal_name")),
        ("Identificación fiscal", issuer.get("tax_id")),
        ("Domicilio", issuer.get("address")),
        ("Origen", doc.get("origin")),
        ("Destino", doc.get("destination")),
        ("Peso bruto total", doc.get("gross_weight_kg")),
        ("Unidad de peso", doc.get("weight_unit")),
        ("Número de bultos", doc.get("packages")),
        ("Total de renglones", doc.get("total_lines")),
    ):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
        c = ws.cell(row=row, column=2, value=value if value is not None else "—")
        if value is None:
            c.fill = _ALERT_FILL
        row += 1

    row += 1
    for i, (nombre, ancho) in enumerate(_COLS, start=1):
        c = ws.cell(row=row, column=i, value=nombre)
        c.font = Font(bold=True, size=9)
        c.fill = _HEADER_FILL
        ws.column_dimensions[c.column_letter].width = ancho
    row += 1

    for line in doc["lines"]:
        for i, key in enumerate(_FIELDS, start=1):
            value = line.get(key)
            c = ws.cell(row=row, column=i, value=value if value is not None else "—")
            c.font = Font(size=9)
            # La celda vacía se ve: es la que alguien tiene que llenar.
            if value is None:
                c.fill = _ALERT_FILL
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
