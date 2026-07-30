"""IFRC packing list Excel (.xlsx) generation.

Columns follow the IFRC/ICRC Emergency Relief Items standard:
  Material code | Description | Unit | Qty | Weight (kg) | Batch | Expiry | Box code | Pallet
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.legal import CUSTOMS_LEGEND_EN, CUSTOMS_LEGEND_ES
from app.utils.manifest import ManifestData

_HEADER_FILL = PatternFill("solid", fgColor="1F5E8C")
_PALLET_FILL = PatternFill("solid", fgColor="EAE1CF")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_PALLET_FONT = Font(bold=True, size=10)
_COLS = [
    ("Código material", 18),
    ("Descripción", 36),
    ("Categoría", 16),
    ("INN / nombre genérico", 22),
    ("Concentración", 14),
    ("Unidad", 10),
    ("Cantidad", 10),
    ("Peso (kg)", 11),
    ("Lote", 14),
    ("Caducidad", 12),
    ("Código caja", 16),
    ("Tarima", 14),
]


def generate_manifest_xlsx(data: ManifestData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List IFRC"

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = f"MANIFIESTO DE ENVÍO · {data.reference or data.shipment_id[:8].upper()}"
    ws["A1"].font = Font(bold=True, size=13)

    meta = [
        ("Destino", data.destination),
        ("Transportista", data.carrier or "—"),
        ("Referencia", data.reference or "—"),
        ("Estado", data.status),
        ("Fecha cierre", data.closed_at.strftime("%d/%m/%Y") if data.closed_at else "—"),
        ("Generado", datetime.now(tz=timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=9)
        ws.cell(row=i, column=2, value=value).font = Font(size=9)

    header_row = len(meta) + 3

    # ── Column headers ────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(_COLS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[header_row].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = header_row + 1
    for pallet in data.pallets:
        # Pallet separator row
        ws.merge_cells(f"A{current_row}:L{current_row}")
        sep = ws.cell(row=current_row, column=1, value=f"Tarima: {pallet.code}")
        sep.font = _PALLET_FONT
        sep.fill = _PALLET_FILL
        sep.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for box in pallet.boxes:
            expiry_str = box.expiry_date.strftime("%d/%m/%Y") if box.expiry_date else ""
            weight_val = float(box.weight_kg) if box.weight_kg is not None else None
            row_data = [
                box.code,                                    # material code (box code serves as ref)
                box.display_name,
                box.category,
                box.inn_name or "",
                box.strength or "",
                box.unit,
                box.quantity,
                weight_val,
                box.batch or "",
                expiry_str,
                box.code,
                pallet.code,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical="center")
                if col_idx in (7, 8):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    total_boxes = sum(len(p.boxes) for p in data.pallets)
    total_units = sum(b.quantity for p in data.pallets for b in p.boxes)
    total_weight = sum(float(b.weight_kg) for p in data.pallets for b in p.boxes if b.weight_kg)

    ws.merge_cells(f"A{current_row}:F{current_row}")
    total_label = ws.cell(row=current_row, column=1, value=f"TOTALES · {total_boxes} cajas")
    total_label.font = Font(bold=True, size=9)

    ws.cell(row=current_row, column=7, value=total_units).font = Font(bold=True, size=9)
    ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=current_row, column=8, value=round(total_weight, 3)).font = Font(bold=True, size=9)
    ws.cell(row=current_row, column=8).alignment = Alignment(horizontal="right")

    # ── Leyenda de aduana ─────────────────────────────────────────────────────
    # Va en la hoja, no en una nota al margen: quien la imprime para aduana
    # imprime la hoja.
    legend_row = current_row + 2
    for offset, texto in enumerate((CUSTOMS_LEGEND_ES, CUSTOMS_LEGEND_EN)):
        fila = legend_row + offset
        ws.merge_cells(f"A{fila}:L{fila}")
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = Font(size=8, italic=offset == 1)
        celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
