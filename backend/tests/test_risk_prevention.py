"""Prevención de riesgos en donaciones en especie (Fase 20, tasks 3 y 10).

Dos controles que se sostienen entre sí:

- **Aceptación registrada.** La donación es una transferencia irrevocable, sin
  derecho a designar destino. Quien dona lo acepta, y queda la versión y la
  fecha de lo que aceptó.
- **Umbral de volumen.** El escrutinio por tipo de donante tiene una evasión
  obvia: registrarse como persona física. A partir de cierto volumen el anonimato
  se acaba, sea quien sea. Es escalamiento, no tope duro: un tope invita a partir
  la donación en pedazos, que es la técnica clásica.

El valor del umbral vive solo en el entorno. Aquí se prueba el mecanismo con
valores de prueba, nunca los operativos.
"""

from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest
from fastapi import HTTPException
from pydantic import ValidationError

from app.legal import CURRENT_DONATION_TERMS_VERSION
from app.utils.volume import exceeds_volume_threshold


# ── Umbral: mecanismo ────────────────────────────────────────────────────────

def test_sin_umbral_configurado_el_control_esta_apagado(monkeypatch):
    """El valor operativo vive en el entorno, no en este repositorio público.

    Sin configurar, no hay umbral que aplicar: el control lo enciende quien
    opera, que es quien sabe qué volumen es normal en su contexto.
    """
    monkeypatch.delenv("DONATION_VOLUME_THRESHOLD_BOXES", raising=False)
    monkeypatch.delenv("DONATION_VOLUME_THRESHOLD_KG", raising=False)
    assert exceeds_volume_threshold(boxes=10_000, kg=10_000) is False


def test_el_umbral_de_cajas_dispara(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "10")
    assert exceeds_volume_threshold(boxes=11, kg=None) is True
    assert exceeds_volume_threshold(boxes=10, kg=None) is False


def test_el_umbral_de_peso_dispara(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_KG", "100")
    assert exceeds_volume_threshold(boxes=1, kg=101) is True
    assert exceeds_volume_threshold(boxes=1, kg=100) is False


def test_cualquiera_de_los_dos_umbrales_basta(monkeypatch):
    """Se mide con lo que el sistema ya tiene. No hay valor comercial que medir,
    y eso es deliberado."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "10")
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_KG", "100")
    assert exceeds_volume_threshold(boxes=1, kg=500) is True
    assert exceeds_volume_threshold(boxes=50, kg=1) is True


def test_un_umbral_mal_escrito_no_tumba_la_captura(monkeypatch):
    """Una variable de entorno con basura no puede impedir que un centro capture."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "muchas")
    assert exceeds_volume_threshold(boxes=99, kg=None) is False


def test_los_valores_operativos_no_viven_en_el_repositorio():
    """El mecanismo se publica; el parámetro que determina cuándo salta, no."""
    from pathlib import Path

    src = Path("app/utils/volume.py").read_text()
    assert "os.environ" in src
    # Sin defaults numéricos: un default sería publicar el umbral.
    assert "DONATION_VOLUME_THRESHOLD_BOXES\", \"" not in src


# ── Umbral en el intake: sobre el umbral no hay anonimato ────────────────────

def _intake_data(donor=None, boxes=1, weight=None):
    data = MagicMock()
    data.donor = donor
    data.donation_id = None
    data.donante_libre = None
    data.notes = None
    data.donor_terms_accepted = True
    data.anonymous_exception_reason = None
    data.boxes = []
    for _ in range(boxes):
        bd = MagicMock()
        bd.product_type_id = uuid4()
        bd.quantity = 1
        bd.unit = "cajas"
        bd.batch = None
        bd.expiry_date = None
        bd.weight_kg = weight
        bd.gtin = None
        data.boxes.append(bd)
    return data


def _correr_intake(data):
    from app.services.intake_service import IntakeService

    db = MagicMock()
    db.get.return_value = None
    with (
        patch("app.services.intake_service.CampaignRepository") as MockCampaign,
        patch("app.services.intake_service.ProductTypeRepository") as MockPt,
        patch("app.services.intake_service.IntakeRepository") as MockIntake,
        patch("app.services.intake_service.UserCampaignRepository") as MockMembership,
        patch("app.services.intake_service.DonorRepository"),
        patch("app.services.intake_service.validate_box", return_value=None),
        patch("app.services.intake_service.IntakeOut"),
        patch("app.services.intake_service.BoxOut"),
        patch("app.services.intake_service.DonorOut"),
    ):
        campaign = MagicMock()
        campaign.is_active = True
        MockCampaign.return_value.find_by_id.return_value = campaign
        MockMembership.return_value.is_member.return_value = True
        MockPt.return_value.find_by_id.return_value = MagicMock(
            category="OTHER", min_shelf_life_days=None
        )
        MockIntake.return_value.save_intake.side_effect = lambda i: i
        IntakeService(db).create(data, uuid4(), uuid4())
        return MockIntake.return_value.save_intake.call_args[0][0]


def test_una_captura_anonima_sobre_el_umbral_exige_identificar_al_donante(monkeypatch):
    """Lo fuerza el sistema, no la memoria de quien captura."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    with pytest.raises(HTTPException) as exc:
        _correr_intake(_intake_data(donor=None, boxes=5))
    assert exc.value.status_code == 400


def test_una_captura_anonima_bajo_el_umbral_sigue_siendo_anonima(monkeypatch):
    """La donación anónima es la norma del dominio y no se toca sin razón."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    intake = _correr_intake(_intake_data(donor=None, boxes=2))
    assert intake.donor_id is None


def test_sobre_el_umbral_con_donante_registrado_la_captura_procede(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    donor = MagicMock(donor_type="fisica")
    intake = _correr_intake(_intake_data(donor=donor, boxes=5))
    assert intake is not None


# ── Aceptación de términos ───────────────────────────────────────────────────

def test_el_pre_registro_exige_aceptar_los_terminos():
    """Sin aceptación no hay transferencia de propiedad que registrar."""
    from app.schemas.donation import DonationCreate

    with pytest.raises(ValidationError):
        DonationCreate(
            donor={"donor_type": "fisica", "first_name": "Ana", "last_name": "P",
                   "email": "ana@ejemplo.test"},
            items=[{"free_text": "3 cobijas", "quantity": 3, "unit": "piezas"}],
            terms_accepted=False,
        )


def test_el_pre_registro_guarda_version_y_fecha_de_lo_aceptado():
    from app.services.donation_service import DonationService

    svc = DonationService(MagicMock())
    data = MagicMock()
    data.donor = MagicMock(email="ana@ejemplo.test")
    data.items = []
    data.notes = None
    data.intended_center_id = None
    data.intended_campaign_id = None

    with (
        patch("app.services.donation_service.DonationRepository") as MockRepo,
        patch("app.services.donation_service.DonorRepository"),
        patch("app.services.donation_service.enqueue"),
    ):
        MockRepo.return_value.find_open_for_email.return_value = None
        donation = svc.submit(data, MagicMock())

    assert donation.terms_version == CURRENT_DONATION_TERMS_VERSION
    assert donation.terms_accepted_at is not None


def test_una_persona_moral_siempre_acepta_los_terminos():
    """Quien dona a nombre de una empresa acepta en el mostrador, sin excepción."""
    from app.services.intake_service import IntakeService  # noqa: F401

    donor = MagicMock(donor_type="moral")
    data = _intake_data(donor=donor)
    data.donor_terms_accepted = False

    with pytest.raises(HTTPException) as exc:
        _correr_intake(data)
    assert exc.value.status_code == 400


def test_el_intake_con_donante_guarda_la_version_aceptada():
    donor = MagicMock(donor_type="fisica")
    intake = _correr_intake(_intake_data(donor=donor))
    assert intake.donor_terms_version == CURRENT_DONATION_TERMS_VERSION


def test_una_captura_anonima_no_registra_aceptacion():
    """No hay a quién atribuírsela: nadie se identificó."""
    intake = _correr_intake(_intake_data(donor=None))
    assert intake.donor_terms_version is None


# ── Volumen atípico en el pre-registro: marca, no bloquea ────────────────────

def test_el_pre_registro_sobre_el_umbral_queda_marcado(monkeypatch):
    """El donante ya viene identificado, así que no se bloquea nada: se marca
    para que el doble check se haga con la guía de banderas rojas a la mano."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    from app.services.donation_service import DonationService

    svc = DonationService(MagicMock())
    data = MagicMock()
    data.donor = MagicMock(email="ana@ejemplo.test")
    data.items = [MagicMock(quantity=10, product_type_id=None, free_text="x", unit="cajas")]
    data.notes = None
    data.intended_center_id = None
    data.intended_campaign_id = None

    with (
        patch("app.services.donation_service.DonationRepository") as MockRepo,
        patch("app.services.donation_service.DonorRepository"),
        patch("app.services.donation_service.enqueue"),
    ):
        MockRepo.return_value.find_open_for_email.return_value = None
        donation = svc.submit(data, MagicMock())

    assert donation.atypical_volume is True


# ── Excepción aprobada: la captura nunca se detiene ──────────────────────────
#
# El umbral pide identificar. Si la persona se niega —que es en sí una bandera
# roja— el voluntario no se queda atorado con el camión en la puerta: registra
# la excepción con motivo, la captura entra, y queda pendiente de que un
# coordinador o national_admin la apruebe o la rechace. Todo con auditoría.

def test_sobre_el_umbral_sin_identificar_y_sin_motivo_no_procede(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    with pytest.raises(HTTPException):
        _correr_intake(_intake_data(donor=None, boxes=5))


def test_con_motivo_de_excepcion_la_captura_entra(monkeypatch):
    """El almacén no se detiene: lo que se abre es una revisión, no un bloqueo."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    data = _intake_data(donor=None, boxes=5)
    data.anonymous_exception_reason = "La persona se negó a identificarse"
    intake = _correr_intake(data)
    assert intake is not None


def test_la_excepcion_abre_una_revision_pendiente(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    from app.models.risk_review import RiskReview

    data = _intake_data(donor=None, boxes=5)
    data.anonymous_exception_reason = "La persona se negó a identificarse"
    revisiones = _revisiones_creadas(data)

    assert len(revisiones) == 1
    assert revisiones[0].kind == "ANONYMOUS_EXCEPTION"
    assert revisiones[0].status == "PENDING"
    assert isinstance(revisiones[0], RiskReview)


def test_sobre_el_umbral_con_donante_tambien_avisa_al_coordinador(monkeypatch):
    """Identificado no quiere decir sin revisar: el volumen atípico se mira."""
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    data = _intake_data(donor=MagicMock(donor_type="fisica"), boxes=5)
    revisiones = _revisiones_creadas(data)

    assert [r.kind for r in revisiones] == ["ATYPICAL_VOLUME"]


def test_una_captura_normal_no_genera_revisiones(monkeypatch):
    monkeypatch.setenv("DONATION_VOLUME_THRESHOLD_BOXES", "3")
    assert _revisiones_creadas(_intake_data(donor=None, boxes=1)) == []


def _revisiones_creadas(data):
    """Las RiskReview que el intake agregó a la sesión."""
    from app.models.risk_review import RiskReview
    from app.services.intake_service import IntakeService

    db = MagicMock()
    db.get.return_value = None
    agregados = []
    db.add.side_effect = lambda obj: agregados.append(obj)

    with (
        patch("app.services.intake_service.CampaignRepository") as MockCampaign,
        patch("app.services.intake_service.ProductTypeRepository") as MockPt,
        patch("app.services.intake_service.IntakeRepository") as MockIntake,
        patch("app.services.intake_service.UserCampaignRepository") as MockMembership,
        patch("app.services.intake_service.DonorRepository"),
        patch("app.services.intake_service.validate_box", return_value=None),
        patch("app.services.intake_service.IntakeOut"),
        patch("app.services.intake_service.BoxOut"),
        patch("app.services.intake_service.DonorOut"),
    ):
        campaign = MagicMock()
        campaign.is_active = True
        MockCampaign.return_value.find_by_id.return_value = campaign
        MockMembership.return_value.is_member.return_value = True
        MockPt.return_value.find_by_id.return_value = MagicMock(
            category="OTHER", min_shelf_life_days=None
        )
        MockIntake.return_value.save_intake.side_effect = lambda i: i
        IntakeService(db).create(data, uuid4(), uuid4())

    return [o for o in agregados if isinstance(o, RiskReview)]


# ── Resolución de la revisión ────────────────────────────────────────────────

def _review_service(revision=None, actor_role="coordinator", actor_id=None):
    from app.services.risk_review_service import RiskReviewService

    db = MagicMock()
    svc = RiskReviewService(db)
    actor = MagicMock(id=actor_id or uuid4(), center_role=actor_role, center_id=uuid4())
    return svc, db, actor


def _revision(center_id=None, created_by=None):
    r = MagicMock()
    r.id = uuid4()
    r.status = "PENDING"
    r.center_id = center_id or uuid4()
    r.created_by_user_id = created_by or uuid4()
    return r


def test_el_coordinador_aprueba_y_queda_quien_y_cuando():
    svc, db, actor = _review_service()
    revision = _revision(center_id=actor.center_id)

    with patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo:
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "APPROVED", "Donante conocido del centro", actor)

    assert revision.status == "APPROVED"
    assert revision.reviewed_by_user_id == actor.id
    assert revision.reviewed_at is not None


def test_rechazar_exige_motivo():
    """Un rechazo sin motivo no le sirve a quien capturó ni a la auditoría."""
    svc, db, actor = _review_service()
    revision = _revision(center_id=actor.center_id)

    with (
        patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo,
        pytest.raises(HTTPException),
    ):
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "REJECTED", "", actor)


def test_un_voluntario_no_puede_resolver():
    """Escalar es el punto: quien captura no se autoriza a sí mismo."""
    svc, db, actor = _review_service(actor_role="volunteer")
    revision = _revision(center_id=actor.center_id)

    with (
        patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo,
        pytest.raises(HTTPException) as exc,
    ):
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "APPROVED", "ok", actor)
    assert exc.value.status_code == 403


def test_nadie_aprueba_su_propia_captura():
    """Separación de funciones: si el coordinador capturó, escala al nacional."""
    svc, db, actor = _review_service()
    revision = _revision(center_id=actor.center_id, created_by=actor.id)

    with (
        patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo,
        pytest.raises(HTTPException) as exc,
    ):
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "APPROVED", "ok", actor)
    assert exc.value.status_code == 403


def test_el_admin_nacional_si_puede_resolver_lo_que_el_mismo_capturo():
    """Es el escalamiento final: si él no puede, la revisión queda muerta."""
    svc, db, actor = _review_service(actor_role="national_admin")
    revision = _revision(created_by=actor.id)

    with patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo:
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "APPROVED", "Revisado con el centro", actor)

    assert revision.status == "APPROVED"


def test_un_coordinador_no_resuelve_revisiones_de_otro_centro():
    svc, db, actor = _review_service()
    revision = _revision()          # otro centro

    with (
        patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo,
        pytest.raises(HTTPException) as exc,
    ):
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "APPROVED", "ok", actor)
    assert exc.value.status_code == 404


def test_una_revision_ya_resuelta_no_se_vuelve_a_resolver():
    svc, db, actor = _review_service()
    revision = _revision(center_id=actor.center_id)
    revision.status = "APPROVED"

    with (
        patch("app.services.risk_review_service.RiskReviewRepository") as MockRepo,
        pytest.raises(HTTPException) as exc,
    ):
        MockRepo.return_value.find_by_id.return_value = revision
        svc.resolve(revision.id, "REJECTED", "no", actor)
    assert exc.value.status_code == 409


def test_la_cola_de_revisiones_esta_montada_y_limitada():
    from pathlib import Path

    src = Path("app/routers/risk_review.py").read_text()
    assert "@limiter.limit" in src
    assert "AuditRepository" in src        # resolver deja rastro fuera de la propia fila
    assert "_NO_CACHE" in src

    main = Path("app/main.py").read_text()
    assert "risk_review.router" in main


# ── Leyenda de aduana (task 4) ───────────────────────────────────────────────
#
# El papel es donde el control rector se vuelve oponible: quien recibe en
# aduana lee que los bienes se transfirieron de forma irrevocable, sin valor
# comercial y sin consignatario designado por quien donó.

def test_la_leyenda_es_bilingue_y_unica():
    """Un solo texto para todos los documentos: si vive en cada plantilla, se
    desfasa en cuanto alguien edite uno."""
    from app.legal import CUSTOMS_LEGEND_EN, CUSTOMS_LEGEND_ES

    assert "irrevocable" in CUSTOMS_LEGEND_ES.lower()
    assert "sin valor comercial" in CUSTOMS_LEGEND_ES.lower()
    assert "irrevocably" in CUSTOMS_LEGEND_EN.lower()


def test_el_manifiesto_de_envio_lleva_la_leyenda():
    from datetime import datetime, timezone

    from app.legal import CUSTOMS_LEGEND_EN, CUSTOMS_LEGEND_ES
    from app.utils.manifest import ManifestData, render_manifest_html

    html = render_manifest_html(ManifestData(
        shipment_id="x", destination="Caracas", carrier=None, reference="EN-0001",
        status="CLOSED", closed_at=datetime.now(timezone.utc),
    ))
    assert CUSTOMS_LEGEND_ES in html
    assert CUSTOMS_LEGEND_EN in html


def test_el_manifiesto_de_transferencia_lleva_la_leyenda():
    from datetime import datetime, timezone

    from app.legal import CUSTOMS_LEGEND_ES
    from app.utils.manifest import TransferManifestData, render_transfer_manifest_html

    html = render_transfer_manifest_html(TransferManifestData(
        transfer_id="x", from_center="A", to_center="B", status="APPROVED",
        created_at=datetime.now(timezone.utc),
    ))
    assert CUSTOMS_LEGEND_ES in html


def test_el_manifiesto_xlsx_lleva_la_leyenda():
    import io

    from openpyxl import load_workbook

    from app.legal import CUSTOMS_LEGEND_ES
    from app.utils.manifest import ManifestData
    from app.utils.manifest_xlsx import generate_manifest_xlsx

    data = ManifestData(
        shipment_id="x", destination="Caracas", carrier=None, reference="EN-0001",
        status="CLOSED", closed_at=None,
    )
    hoja = load_workbook(io.BytesIO(generate_manifest_xlsx(data))).active
    textos = [c.value for row in hoja.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(CUSTOMS_LEGEND_ES in t for t in textos)


def test_la_etiqueta_de_tarima_lleva_la_leyenda():
    """La tarima viaja sola: su etiqueta tiene que sostener la declaración."""
    from pathlib import Path

    src = Path("app/utils/pdf_pallet_label.py").read_text()
    assert "CUSTOMS_LEGEND" in src


def test_una_tarima_con_muchas_cajas_no_tapa_la_leyenda():
    """Antes el listado escribía encima de la leyenda y hasta fuera de la hoja."""
    from reportlab.lib.units import mm

    from app.utils.pdf_pallet_label import _LEGEND_TOP, fit_box_codes

    codigos = [f"BX-{i:04d}" for i in range(300)]
    visibles, restantes = fit_box_codes(codigos, y_start=162 * mm)

    assert restantes > 0                       # con 300 cajas algo se queda fuera
    assert len(visibles) + restantes == len(codigos)
    # La última fila dibujada no baja de donde empieza la leyenda.
    assert 162 * mm - (len(visibles) // 3) * 5 * mm >= _LEGEND_TOP


def test_una_tarima_chica_muestra_todas_sus_cajas():
    from reportlab.lib.units import mm

    from app.utils.pdf_pallet_label import fit_box_codes

    codigos = [f"BX-{i:04d}" for i in range(12)]
    assert fit_box_codes(codigos, y_start=162 * mm) == (codigos, 0)


def test_la_etiqueta_se_genera_con_una_tarima_enorme():
    """Prueba de humo: el recorte no puede romper la generación."""
    from app.utils.pdf_pallet_label import PalletLabelData, generate_pallet_label_pdf

    pdf = generate_pallet_label_pdf(PalletLabelData(
        code="TM-GRANDE", center_name="Centro A", status="CLOSED",
        box_codes=[f"BX-{i:04d}" for i in range(300)],
    ))
    assert pdf.startswith(b"%PDF")


# ── Guía de banderas rojas (task 5) ──────────────────────────────────────────

def test_la_guia_de_banderas_rojas_existe_en_ambos_idiomas():
    """Publicar tipologías es deliberado: FATF y los bancos publican las suyas,
    y el valor disuasorio supera al riesgo."""
    from pathlib import Path

    raiz = Path(__file__).resolve().parents[2] / "frontend" / "content" / "manuals"
    for ruta in (raiz / "banderas-rojas.html", raiz / "en" / "banderas-rojas.html"):
        assert ruta.exists(), ruta
        texto = ruta.read_text()
        assert len(texto) > 2000


def test_la_guia_no_publica_el_umbral():
    """El mecanismo se explica; el valor que lo dispara vive en el entorno."""
    import re
    from pathlib import Path

    raiz = Path(__file__).resolve().parents[2] / "frontend" / "content" / "manuals"
    for ruta in (raiz / "banderas-rojas.html", raiz / "en" / "banderas-rojas.html"):
        texto = ruta.read_text()
        assert "DONATION_VOLUME_THRESHOLD" not in texto
        # Sin cifras de cajas o kilos que revelen dónde salta el control.
        assert not re.search(r"\b\d+\s*(cajas|kg|boxes)\b", texto, re.I)


def test_la_guia_esta_registrada_en_el_indice_de_ayuda():
    from pathlib import Path

    registro = (Path(__file__).resolve().parents[2] / "frontend" / "app" / "dashboard"
                / "ayuda" / "manuals.ts").read_text()
    assert "banderas-rojas" in registro
